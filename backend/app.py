# -*- coding: utf-8 -*-
"""
HC Forecast Board · V1 后端底座
- FastAPI + SQLite（本地开发库；企业版数据库批下后仅替换本存储层）
- 高压线：所有数字须来自真实数据源/人工录入；库中无数据一律返回空（不编造、不补零）
- 校验闸：已发生月锁定、备注必填、只读账号拒写、导入整批校验不合格拒绝入库
- 登录为 iOA 占位（X-User 头，默认 bonniewbli）；正式版接 iOA 统一登录
启动：uvicorn app:app --port 8787（工作目录 backend/）
"""
import csv
import io
import json
import os
import random
import sqlite3
import time
from contextlib import contextmanager
from typing import Optional

from fastapi import FastAPI, File, Header, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, PlainTextResponse
from pydantic import BaseModel

DB_PATH = os.environ.get("HCFB_DB") or os.path.join(os.path.dirname(__file__), "hcfb.db")
FRONT_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))  # 仓库根（index.html/admin.html）

app = FastAPI(title="HC Forecast Board API", version="0.1.0")
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])

# ---------------- 模块装配（按看板分工拆分；行为与单文件版完全一致） ----------------
# meta=指标口径 · store=存储/权限/审计 · calc_kb1=看板1/2运算引擎 · sources=外部源直连 · kb3_ledger=看板3台账解析
from meta import (CANON_PROJECTS, OUT_KEYS, CAMP_KEYS, IN_DIRECT_KEYS, BP_EDITABLE,
                  IMPORTABLE, VALUE_ABS_MAX, PLAN_METRICS, PLAN_BRANCH_SECS, EXTRA_METRICS, NAT_N_DEFAULT)
from store import (DB_PATH, db, init_db, now, _audit, _write_cell, get_account,
                   require_writer, require_admin, _grid, _branches)
from calc_kb1 import compute
from sources import SOURCE_METRICS, load_sources_cfg, fetch_source, _month_completed
from kb3_ledger import (LEDGER_CLS, LEDGER_DATE_F, LEDGER_F2DB, LEDGER_REQUIRED, LEDGER_ST_CANON,
                        parse_ledger, validate_ledger_rows, _ledger_list, _ledger_norm, _norm_date, _norm_status)

init_db()


# ---------------- 通用 ----------------
@app.get("/api/health")
def health():
    return {"ok": True, "ts": now(), "storage": "sqlite-dev（企业版数据库批后替换）", "version": app.version}


# ---------------- 配置（与前端 localStorage cfg 同构） ----------------
@app.get("/api/config")
def get_config():
    with db() as c:
        projs = [
            {"key": r["key"], "sec": r["sec"], "name": r["name"], "src": r["src"], "srcCls": r["src_cls"],
             "add": bool(r["add_ok"]), "unbind": bool(r["unbind"]), "on": bool(r["on_ok"]), "sys": bool(r["sys"])}
            for r in c.execute("SELECT * FROM projects ORDER BY pos")
        ]
        accts = [
            {"id": r["id"], "name": r["name"], "role": r["role"], "dept": r["dept"],
             "kb": json.loads(r["kb"] or "[1,1,1,1]"), "on": bool(r["on_ok"]), "demo": bool(r["demo"])}
            for r in c.execute("SELECT * FROM accounts")
        ]
        return {"projs": projs, "accts": accts, "ts": int(time.time() * 1000)}


class ConfigDoc(BaseModel):
    projs: list
    accts: list


@app.put("/api/config")
def put_config(doc: ConfigDoc, x_user: str = Header("bonniewbli")):
    with db() as c:
        require_admin(c, x_user)
        c.execute("DELETE FROM projects")
        for i, p in enumerate(doc.projs):
            c.execute(
                "INSERT INTO projects(key,sec,name,src,src_cls,add_ok,unbind,on_ok,sys,pos) VALUES(?,?,?,?,?,?,?,?,?,?)",
                (p.get("key") or f"x{int(time.time()*1000)}_{i}", p.get("sec", ""), p.get("name", ""),
                 p.get("src", ""), p.get("srcCls", "bp"), int(bool(p.get("add"))), int(bool(p.get("unbind"))),
                 int(p.get("on", True)), int(bool(p.get("sys"))), i),
            )
        c.execute("DELETE FROM accounts")
        for a in doc.accts:
            c.execute(
                "INSERT INTO accounts(id,name,role,dept,kb,on_ok,demo) VALUES(?,?,?,?,?,?,?)",
                (a["id"], a.get("name", ""), a.get("role", "HRBP·可编辑"), a.get("dept", ""),
                 json.dumps(a.get("kb", [1, 1, 1, 1])), int(a.get("on", True)), int(bool(a.get("demo")))),
            )
        _audit(c, x_user, "配置更新", f"项目 {len(doc.projs)} 项 / 账号 {len(doc.accts)} 个（管理后台下发）")
        return {"ok": True}


# ---------------- 年份 ----------------
@app.get("/api/years")
def list_years():
    with db() as c:
        return [dict(r) for r in c.execute("SELECT * FROM years ORDER BY year")]


class YearNew(BaseModel):
    year: int


@app.post("/api/years")
def add_year(y: YearNew, x_user: str = Header("bonniewbli")):
    if y.year < 2000 or y.year > 2100:
        raise HTTPException(422, "年份无效")
    with db() as c:
        require_writer(c, x_user)
        if c.execute("SELECT 1 FROM years WHERE year=?", (y.year,)).fetchone():
            raise HTTPException(409, "年份已存在")
        c.execute("INSERT INTO years(year,status,lock_month) VALUES(?,?,0)", (y.year, "待接入·待编制"))
        _audit(c, x_user, "新增年份", f"{y.year}：生成空模板（看板0-3 · 独立数据空间）")
        _audit(c, "system", "读取系统数", f"{y.year}：各数据源 API 未接入 → 单元格留空待取数（不编造）")
        return {"ok": True}


# ---------------- 看板读写 ----------------
@app.get("/api/board/{year}")
def get_board(year: int):
    with db() as c:
        yr = c.execute("SELECT * FROM years WHERE year=?", (year,)).fetchone()
        if not yr:
            raise HTTPException(404, "年份不存在")
        vals, notes = _grid(c, year)
        brs = _branches(c, year)
        prev_er = _grid(c, year - 1)[0].get("er_out")
        nat_n = yr["nat_n"] if "nat_n" in yr.keys() else NAT_N_DEFAULT
        comp = compute(vals, brs, yr["lock_month"], prev_er=prev_er, nat_n=nat_n)
        metrics = {k: {"vals": vals.get(k, [None] * 12), "notes": notes.get(k, {})} for k, *_ in CANON_PROJECTS}
        for k in EXTRA_METRICS:  # 看板2 期初基线（fa_hc/q_init）一并下发
            metrics[k] = {"vals": vals.get(k, [None] * 12), "notes": notes.get(k, {})}
        metrics["o_nat"]["vals"] = comp["o_nat_eff"]  # 存量优先，派生只补未发生月空格
        # 260723 口径：链行并入实际行——未发生月空格由预估链补（月末实际在岗/期末在岗预估）
        av = list(metrics["actual"]["vals"])
        for m in range(yr["lock_month"], 12):
            if av[m] is None and comp["chain"][m] is not None:
                av[m] = comp["chain"][m]
        metrics["actual"]["vals"] = av
        demo = bool(c.execute("SELECT 1 FROM cells WHERE year=? AND source='demo' LIMIT 1", (year,)).fetchone()
                    or c.execute("SELECT 1 FROM branches WHERE year=? AND created_by='demo' LIMIT 1", (year,)).fetchone()
                    or c.execute("SELECT 1 FROM ledger_rows WHERE batch=-999 LIMIT 1").fetchone())
        return {"year": year, "status": yr["status"], "lock": yr["lock_month"],
                "metrics": metrics, "branches": brs, "computed": comp, "nat": comp["nat"],
                "demo": demo, "ts": int(time.time() * 1000)}


class CellEdit(BaseModel):
    metric: str
    month: int  # 1-12
    value: Optional[float] = None  # None=清空
    note: str = ""  # 260723 起可选：Excel 式直填不强制备注，右键可补


@app.post("/api/board/{year}/cell")
def edit_cell(year: int, e: CellEdit, x_user: str = Header("bonniewbli")):
    with db() as c:
        require_writer(c, x_user)
        yr = c.execute("SELECT * FROM years WHERE year=?", (year,)).fetchone()
        if not yr:
            raise HTTPException(404, "年份不存在")
        if not (1 <= e.month <= 12):
            raise HTTPException(422, "月份须为 1-12")
        base_metric = e.metric.split(":", 1)[0] if e.metric.startswith("branch") else e.metric
        locked = base_metric not in PLAN_METRICS and e.month <= yr["lock_month"]
        if locked and e.metric.startswith("branch:"):
            pb = c.execute("SELECT sec FROM branches WHERE id=?", (int(e.metric.split(":", 1)[1]),)).fetchone()
            if pb and pb["sec"] in PLAN_BRANCH_SECS:
                locked = False  # 看板2 计划类分支（法定HC/预算当量·其中）：规划数据，已发生月可改
        if locked:
            raise HTTPException(423, f"{e.month}月为已发生月（已锁定），改动须走「采纳修正」流程")
        # 260723 交互调整：备注改为可选（Excel 式直填；右键单元格可补备注）——改动本身仍全量审计留痕
        if e.value is not None and abs(e.value) > VALUE_ABS_MAX:
            raise HTTPException(422, "量级异常，拒绝入库")
        if e.metric.startswith("branch:"):
            bid = int(e.metric.split(":", 1)[1])
            b = c.execute("SELECT * FROM branches WHERE id=? AND year=?", (bid, year)).fetchone()
            if not b:
                raise HTTPException(404, "分支不存在")
            c.execute(
                "INSERT INTO branch_cells(branch_id,month,value,note,updated_by,updated_at) VALUES(?,?,?,?,?,?) "
                "ON CONFLICT(branch_id,month) DO UPDATE SET value=excluded.value,note=excluded.note,"
                "updated_by=excluded.updated_by,updated_at=excluded.updated_at",
                (bid, e.month, e.value, e.note.strip(), x_user, now()),
            )
            _audit(c, x_user, "调节录入", f"{year} 分支「{b['name']}」 {e.month}月 → {e.value}（{e.note.strip()}）")
        elif e.metric in EXTRA_METRICS:
            _write_cell(c, year, e.metric, e.month, e.value, e.note.strip(), "bp", x_user)
            _audit(c, x_user, "调节录入", f"{year}「{EXTRA_METRICS[e.metric]}」 {e.month}月 → {e.value}（{e.note.strip()}）")
        else:
            p = c.execute("SELECT * FROM projects WHERE key=?", (e.metric,)).fetchone()
            if not p:
                raise HTTPException(404, "指标不存在")
            if e.metric not in BP_EDITABLE:
                raise HTTPException(403, f"「{p['name']}」为系统数指标，不可手工录入（走数据源/上传兜底）")
            if not p["add_ok"]:
                raise HTTPException(403, f"「{p['name']}」已在管理后台关闭手动录入")
            _write_cell(c, year, e.metric, e.month, e.value, e.note.strip(), "bp", x_user)
            _audit(c, x_user, "调节录入", f"{year}「{p['name']}」 {e.month}月 → {e.value}（{e.note.strip()}）")
    return get_board(year)


class BranchNew(BaseModel):
    sec: str
    name: str
    sign: str  # '+' / '-'


@app.post("/api/board/{year}/branch")
def add_branch(year: int, b: BranchNew, x_user: str = Header("bonniewbli")):
    if b.sign not in ("+", "-"):
        raise HTTPException(422, "方向须为 + 或 −")
    if not b.name.strip():
        raise HTTPException(422, "分支名称必填")
    with db() as c:
        require_writer(c, x_user)
        c.execute(
            "INSERT INTO branches(year,sec,name,sign,on_ok,created_by,created_at) VALUES(?,?,?,?,1,?,?)",
            (year, b.sec, b.name.strip(), b.sign, x_user, now()),
        )
        _audit(c, x_user, "新增分支", f"{year} {b.sec} · {b.name.strip()}（{b.sign}）")
    return get_board(year)


@app.delete("/api/board/{year}/branch/{bid}")
def del_branch(year: int, bid: int, x_user: str = Header("bonniewbli")):
    with db() as c:
        require_writer(c, x_user)
        b = c.execute("SELECT * FROM branches WHERE id=? AND year=?", (bid, year)).fetchone()
        if not b:
            raise HTTPException(404, "分支不存在")
        c.execute("DELETE FROM branches WHERE id=?", (bid,))
        _audit(c, x_user, "删除分支", f"{year} {b['sec']} · {b['name']}")
    return get_board(year)


# ---------------- 上传兜底：CSV 导入 → 校验闸 → 快照入库 ----------------
@app.post("/api/import/{year}")
async def import_csv(year: int, file: UploadFile = File(...), x_user: str = Header("bonniewbli")):
    with db() as c:
        require_writer(c, x_user)
        if not c.execute("SELECT 1 FROM years WHERE year=?", (year,)).fetchone():
            raise HTTPException(404, "年份不存在")
    raw = await file.read()
    try:
        text = raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = raw.decode("gbk", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows, errors = [], []
    header_skipped = False
    for ln, row in enumerate(reader, start=1):
        if not row or all(not x.strip() for x in row):
            continue
        if len(row) < 3:
            errors.append(f"第{ln}行：缺字段（需 metric,month,value）")
            continue
        metric, month_s, value_s = row[0].strip(), row[1].strip(), row[2].strip()
        if not header_skipped and metric.lower() in ("metric", "指标", "key"):
            header_skipped = True
            continue
        if metric not in IMPORTABLE:
            errors.append(f"第{ln}行：未知/不可导入指标「{metric}」（可导入：{','.join(sorted(IMPORTABLE))}）")
            continue
        if not month_s.isdigit() or not (1 <= int(month_s) <= 12):
            errors.append(f"第{ln}行：期间错（month 须 1-12，收到「{month_s}」）")
            continue
        try:
            value = float(value_s)
        except ValueError:
            errors.append(f"第{ln}行：数值无效「{value_s}」")
            continue
        if abs(value) > VALUE_ABS_MAX:
            errors.append(f"第{ln}行：量级异常（|{value}|>{VALUE_ABS_MAX}）")
            continue
        rows.append((metric, int(month_s), value))
    if errors:
        raise HTTPException(422, {"msg": "数据不合格，整批拒绝入库", "errors": errors[:50], "total_errors": len(errors)})
    if not rows:
        raise HTTPException(422, "文件无有效数据行")
    with db() as c:
        require_writer(c, x_user)
        yr = c.execute("SELECT * FROM years WHERE year=?", (year,)).fetchone()
        lock = yr["lock_month"]
        c.execute(
            "INSERT INTO snapshots(year,filename,rows_n,created_by,created_at) VALUES(?,?,?,?,?)",
            (year, file.filename, len(rows), x_user, now()),
        )
        snap_id = c.execute("SELECT last_insert_rowid() AS i").fetchone()["i"]
        applied, diffs = 0, 0
        for metric, month, value in rows:
            cur = c.execute("SELECT value FROM cells WHERE year=? AND metric=? AND month=?", (year, metric, month)).fetchone()
            curv = cur["value"] if cur else None
            confirmed = metric not in PLAN_METRICS and month <= lock and curv is not None
            if confirmed and curv != value:
                # 迟到数据：已确认月不自动改——留差异，待「采纳修正/维持口径」
                ex = c.execute(
                    "SELECT id FROM pending_diffs WHERE year=? AND metric=? AND month=? AND status='open'",
                    (year, metric, month)).fetchone()
                if ex:
                    c.execute("UPDATE pending_diffs SET src_value=?, source=?, created_at=? WHERE id=?",
                              (value, f"import#{snap_id}", now(), ex["id"]))
                else:
                    c.execute(
                        "INSERT INTO pending_diffs(year,metric,month,cur_value,src_value,source,status,created_at) "
                        "VALUES(?,?,?,?,?,?,'open',?)",
                        (year, metric, month, curv, value, f"import#{snap_id}", now()))
                diffs += 1
            else:
                _write_cell(c, year, metric, month, value, None, f"import#{snap_id}", x_user)
                applied += 1
        _audit(c, x_user, "导入快照",
               f"{year} 批次#{snap_id}「{file.filename}」采纳 {applied} 格" +
               (f"；已确认月差异 {diffs} 格待处理（源数据已变化）" if diffs else "，校验闸通过入库"))
    return {"ok": True, "snapshot": snap_id, "rows": len(rows), "applied": applied, "diffs": diffs}


@app.get("/api/sources")
def sources_status():
    cfg = load_sources_cfg()
    out = []
    for k, name in SOURCE_METRICS.items():
        entries = cfg.get(k) or []
        if isinstance(entries, dict):
            entries = [entries]
        out.append({"metric": k, "name": name,
                    "configured": bool(entries and entries[0].get("url")),
                    "sources": [{"name": e.get("name", "?"), "url": e.get("url", ""), "note": e.get("note", "")} for e in entries]})
    return {"sources": out, "cfg_path": "backend/sources_config.json（gitignore·凭据不进仓库）"}


class SyncReq(BaseModel):
    year: int
    source: Optional[str] = None  # 多来源时指定 name（如 diy 第二来源核对），缺省用第一个


@app.post("/api/sources/{metric}/sync")
def source_sync(metric: str, q: SyncReq, x_user: str = Header("bonniewbli")):
    if metric not in SOURCE_METRICS or metric not in IMPORTABLE:
        raise HTTPException(404, f"指标「{metric}」不支持外部源直连")
    entries = load_sources_cfg().get(metric) or []
    if isinstance(entries, dict):
        entries = [entries]
    cfg = next((e for e in entries if not q.source or e.get("name") == q.source), None)
    if not cfg or not cfg.get("url"):
        raise HTTPException(428, {"msg": f"「{SOURCE_METRICS[metric]}」外部源未配置",
                                  "need": "backend/sources_config.json 填入该指标的 url/method/headers/params/map（参照 sources_config.example.json；来源=F12抓包 Copy as cURL 或平台开放API文档）"})
    with db() as c:
        require_writer(c, x_user)
        yr = c.execute("SELECT * FROM years WHERE year=?", (q.year,)).fetchone()
        if not yr:
            raise HTTPException(404, "年份不存在")
    months_vals = fetch_source(cfg, q.year)
    bad = [f"{m}月={v}" for m, v in months_vals.items() if abs(v) > VALUE_ABS_MAX]
    if bad:
        raise HTTPException(422, {"msg": "量级异常，整批拒绝入库", "errors": bad})
    with db() as c:
        require_writer(c, x_user)
        lock = yr["lock_month"]
        src_tag = f"sync:{cfg.get('name', metric)}"
        applied, diffs, skipped = 0, 0, []
        for m in sorted(months_vals):
            v = months_vals[m]
            if not _month_completed(q.year, m):
                skipped.append(m)  # 未完结月：月末快照尚不存在
                continue
            cur = c.execute("SELECT value FROM cells WHERE year=? AND metric=? AND month=?", (q.year, metric, m)).fetchone()
            curv = cur["value"] if cur else None
            confirmed = metric not in PLAN_METRICS and m <= lock and curv is not None
            if confirmed and curv != v:
                ex = c.execute("SELECT id FROM pending_diffs WHERE year=? AND metric=? AND month=? AND status='open'",
                               (q.year, metric, m)).fetchone()
                if ex:
                    c.execute("UPDATE pending_diffs SET src_value=?, source=?, created_at=? WHERE id=?",
                              (v, src_tag, now(), ex["id"]))
                else:
                    c.execute("INSERT INTO pending_diffs(year,metric,month,cur_value,src_value,source,status,created_at) "
                              "VALUES(?,?,?,?,?,?,'open',?)", (q.year, metric, m, curv, v, src_tag, now()))
                diffs += 1
            else:
                _write_cell(c, q.year, metric, m, v, None, src_tag, x_user)
                applied += 1
        _audit(c, x_user, "源同步",
               f"{q.year}「{SOURCE_METRICS[metric]}」← {cfg.get('name', '?')}：采纳 {applied} 格" +
               (f"；已确认月差异 {diffs} 格待处理" if diffs else "") +
               (f"；跳过未完结月 {','.join(map(str, skipped))}（月末快照未成立）" if skipped else ""))
    return {"ok": True, "metric": metric, "source": cfg.get("name", "?"),
            "applied": applied, "diffs": diffs, "skipped": skipped}


# ---------------- 示例（演示假数）数据：source='demo' 全程打标，一键彻底清除 ----------------
# 高压线兜底：示例数据只填当前为空的格（绝不覆盖真实数据）；页面挂【示例】横幅；
# 清除 = 按 demo 标签删 cells/branch/台账/历史，真实数据分毫不动。
DEMO_LEDGER_BATCH = -999


@app.post("/api/demo/load")
def demo_load(y: YearNew, x_user: str = Header("bonniewbli")):
    """一次填所有年份页签（含历史归档年）：按各年 lock 填历史实际月，只填空格不覆盖真数"""
    with db() as c:
        require_writer(c, x_user)
        if not c.execute("SELECT 1 FROM years WHERE year=?", (y.year,)).fetchone():
            raise HTTPException(404, "年份不存在")
        filled, skipped = 0, 0
        for yr in c.execute("SELECT * FROM years ORDER BY year").fetchall():
            yy, lock = yr["year"], yr["lock_month"]
            base = 548 + (2026 - yy) * 6  # 历史年在岗基数略高，逐年递减更像真实走势

            def put(metric, month, value, note=None):
                nonlocal filled, skipped
                if not (1 <= month <= 12):
                    return
                if c.execute("SELECT 1 FROM cells WHERE year=? AND metric=? AND month=?", (yy, metric, month)).fetchone():
                    skipped += 1  # 已有数据（真实或旧示例）不覆盖
                    return
                c.execute("INSERT INTO cells(year,metric,month,value,note,source,updated_by,updated_at) VALUES(?,?,?,?,?,?,?,?)",
                          (yy, metric, month, value, note, "demo", x_user, now()))
                filled += 1

            for m in range(1, 13):
                put("budget", m, 550 + (2026 - yy) * 5)
                put("fa_hc", m, 479)
                put("q_init", m, 550 + (2026 - yy) * 5)
            # 历史实际月：只随机填「源」单元格（Excel 思路——运算行 o_nat/总流出/总流入/链 由引擎公式现算，不落库）
            rng = random.Random(yy * 97 + 7)  # 每年固定种子：重复导入结果一致（幂等）
            for m in range(1, lock + 1):
                put("actual", m, base - m + rng.randint(-2, 2))  # 月末快照·随机波动
                put("er_out", m, rng.randint(4, 9))  # ER实际离职：o_nat 已发生月由引擎从此带出，未发生月=近n月均值
                put("o_sys", m, rng.randint(1, 4))
                put("i_soc", m, rng.randint(1, 5))
                if m % 3 == 2:
                    put("o_bp", m, 1, "【示例】历史月已明确离职1人")
                if m % 4 == 0:
                    put("o_act", m, 1, "【示例】历史月计划优化1人")
                if m % 6 == 4:
                    put("i_incr", m, 1, "【示例】历史月增量补位1人")
            if lock >= 2:
                put("i_yy", 2, 3, "【示例】春季批次到岗")
            if lock >= 3:
                put("i_bs", 3, 2, "【示例】毕业生转聘2人")
            if lock >= 5:
                put("i_cbp", 5, 1, "【示例】BP补录1人")
            if lock >= 7:
                put("i_yy", 7, 22, "【示例】历史校招批次")
            if lock < 12:  # 执行中/规划年：未发生月的 BP 调节与分支
                put("o_sys", lock + 1, 2)
                put("o_bp", lock + 3, 3, "【示例】某中心已明确离职3人")
                put("o_act", lock + 4, 2, "【示例】计划优化2人")
                put("i_soc", lock + 1, 4)
                put("i_soc", lock + 2, 3)
                put("i_soc", lock + 3, 2)
                put("i_yy", lock + 1, 25, "【示例】校招批次到岗")
                put("i_bs", lock + 2, 5)
                put("i_cbp", lock + 3, 1, "【示例】BP补录1人")
                brow = c.execute("SELECT id FROM branches WHERE year=? AND created_by='demo'", (yy,)).fetchone()
                if brow:
                    bid = brow["id"]
                else:
                    c.execute("INSERT INTO branches(year,sec,name,sign,on_ok,created_by,created_at) VALUES(?,?,?,?,1,'demo',?)",
                              (yy, "总流出（−）", "【示例】组织架构腾挪", "+", now()))
                    bid = c.execute("SELECT last_insert_rowid() AS i").fetchone()["i"]
                # 260723 起分支纯计人头：负数=调减（示范系统数冲抵写法）
                for bm, bv, bnote in [(lock + 3, -2, "【示例】腾挪冲抵-2（负数=调减）"), (max(lock - 2, 1), 1, "【示例】历史月腾挪1")]:
                    if 1 <= bm <= 12:
                        c.execute("INSERT OR IGNORE INTO branch_cells(branch_id,month,value,note,updated_by,updated_at) VALUES(?,?,?,?,'demo',?)",
                                  (bid, bm, bv, bnote, now()))
        if not c.execute("SELECT 1 FROM ledger_rows WHERE batch=?", (DEMO_LEDGER_BATCH,)).fetchone():
            mm = lambda off: f"{y.year}-{min(lock + off, 12):02d}-15"
            demo_rows = [
                ("已入职", "", mm(0), "【示例】张三", "后台开发", mm(0)),
                ("已offer待入职", mm(1), "", "【示例】李四", "产品经理", ""),
                ("简历&面试中", mm(2), "", "【示例】王五", "前端开发", ""),
                ("Hold", "", "", "【示例】赵六", "测试开发", ""),
            ]
            for st, eta, join_dt, who, job, jd in demo_rows:
                c.execute("INSERT INTO ledger_rows(year,batch,dept,center,owner,src,job,lvl,cls,loc,ask,num,tgt,st,eta,memo,offer,olvl,join_dt,jmemo,who) "
                          "VALUES(0,?,?,?,?,?,?,'','','深圳',?,'1','',?,?,?,?,'',?,'',?)",
                          (DEMO_LEDGER_BATCH, "【示例】云产品五部", "【示例】某中心", "【示例】负责人", "【示例】演示行",
                           job, f"{y.year}-{max(lock - 1, 1):02d}-01", st, eta, "【示例】", who, join_dt or jd, who))
        _audit(c, x_user, "导入示例数据",
               f"全部年份页签（含历史归档年填满实际月）：示例(demo标签)填充 {filled} 格（跳过已有数据 {skipped} 格·不覆盖真实数）+ 示例分支/台账4行；页面挂【示例】横幅，说「删除假数」一键全清")
    return get_board(y.year)


@app.post("/api/demo/clear")
def demo_clear(x_user: str = Header("bonniewbli")):
    with db() as c:
        require_writer(c, x_user)
        n_cells = c.execute("SELECT COUNT(*) AS n FROM cells WHERE source='demo'").fetchone()["n"]
        c.execute("DELETE FROM cells WHERE source='demo'")
        c.execute("DELETE FROM cells_history WHERE source='demo'")
        bids = [r["id"] for r in c.execute("SELECT id FROM branches WHERE created_by='demo'")]
        for bid in bids:
            c.execute("DELETE FROM branch_cells WHERE branch_id=?", (bid,))
            c.execute("DELETE FROM branches WHERE id=?", (bid,))
        n_led = c.execute("SELECT COUNT(*) AS n FROM ledger_rows WHERE batch=?", (DEMO_LEDGER_BATCH,)).fetchone()["n"]
        c.execute("DELETE FROM ledger_rows WHERE batch=?", (DEMO_LEDGER_BATCH,))
        c.execute("DELETE FROM pending_diffs WHERE source='demo'")
        _audit(c, x_user, "清除示例数据",
               f"按 demo 标签彻底清除：{n_cells} 格 + 分支 {len(bids)} 条 + 台账 {n_led} 行（真实数据不动，审计留痕）")
    return {"ok": True, "cells": n_cells, "branches": len(bids), "ledger": n_led}


# ---------------- 审计 / 导出 ----------------
@app.get("/api/audit")
def get_audit(limit: int = 200):
    with db() as c:
        return [dict(r) for r in c.execute("SELECT * FROM audit ORDER BY id DESC LIMIT ?", (min(limit, 1000),))]


@app.get("/api/export/{year}.csv")
def export_csv(year: int):
    board = get_board(year)
    names = {k: n for k, _s, n, *_ in CANON_PROJECTS}
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["项目"] + [f"{m}月" for m in range(1, 13)] + ["年均"])
    def fmt(v):
        return "" if v is None else (int(v) if float(v).is_integer() else v)
    w.writerow([names["budget"]] + [fmt(v) for v in board["metrics"]["budget"]["vals"]] + [""])
    w.writerow(["总流出（−）"] + [fmt(v) for v in board["computed"]["outT"]] + [""])
    w.writerow(["总流入（＋）"] + [fmt(v) for v in board["computed"]["inT"]] + [""])
    # 260723 口径：实际与预估合一行（已发生月=实际，未发生月=链）
    w.writerow([names["actual"]] + [fmt(v) for v in board["metrics"]["actual"]["vals"]] + [board["computed"]["chain_avg"] or ""])
    with db() as c:
        _audit(c, "system", "导出", f"{year} 看板1 导出 CSV")
    return PlainTextResponse("﻿" + buf.getvalue(), media_type="text/csv; charset=utf-8")


# ---------------- 迟到数据差异：采纳修正 / 维持口径 ----------------
def _metric_name(c, key):
    if key == "er_out":
        return "ER报表·月实际离职数（HR数仓·URL待求）"
    if key in EXTRA_METRICS:
        return EXTRA_METRICS[key]
    p = c.execute("SELECT name FROM projects WHERE key=?", (key,)).fetchone()
    return p["name"] if p else key


@app.get("/api/diffs/{year}")
def get_diffs(year: int):
    with db() as c:
        out = []
        for r in c.execute("SELECT * FROM pending_diffs WHERE year=? AND status='open' ORDER BY month,metric", (year,)):
            d = dict(r)
            d["metric_name"] = _metric_name(c, d["metric"])
            out.append(d)
        return {"year": year, "diffs": out}


@app.post("/api/diffs/{did}/accept")
def diff_accept(did: int, x_user: str = Header("bonniewbli")):
    with db() as c:
        require_writer(c, x_user)
        d = c.execute("SELECT * FROM pending_diffs WHERE id=? AND status='open'", (did,)).fetchone()
        if not d:
            raise HTTPException(404, "差异不存在或已处理")
        _write_cell(c, d["year"], d["metric"], d["month"], d["src_value"], "采纳修正（迟到数据）", d["source"], x_user)
        c.execute("UPDATE pending_diffs SET status='accepted', resolved_by=?, resolved_at=? WHERE id=?", (x_user, now(), did))
        _audit(c, x_user, "采纳修正",
               f"{d['year']}「{_metric_name(c, d['metric'])}」{d['month']}月：{d['cur_value']} → {d['src_value']}（预估链重锚重算·校验重跑）")
    return get_board(d["year"])


@app.post("/api/diffs/{did}/keep")
def diff_keep(did: int, x_user: str = Header("bonniewbli")):
    with db() as c:
        require_writer(c, x_user)
        d = c.execute("SELECT * FROM pending_diffs WHERE id=? AND status='open'", (did,)).fetchone()
        if not d:
            raise HTTPException(404, "差异不存在或已处理")
        c.execute("UPDATE pending_diffs SET status='kept', resolved_by=?, resolved_at=? WHERE id=?", (x_user, now(), did))
        _audit(c, x_user, "维持口径",
               f"{d['year']}「{_metric_name(c, d['metric'])}」{d['month']}月差异（{d['cur_value']} vs 源 {d['src_value']}）留案，随下周期滚入")
        return {"ok": True}


# ---------------- 月份确认：锁定推进 ----------------
class LockSet(BaseModel):
    lock_month: int


@app.post("/api/years/{year}/lock")
def set_lock(year: int, s: LockSet, x_user: str = Header("bonniewbli")):
    if not (0 <= s.lock_month <= 12):
        raise HTTPException(422, "锁定月须为 0-12")
    with db() as c:
        require_writer(c, x_user)
        yr = c.execute("SELECT * FROM years WHERE year=?", (year,)).fetchone()
        if not yr:
            raise HTTPException(404, "年份不存在")
        c.execute("UPDATE years SET lock_month=? WHERE year=?", (s.lock_month, year))
        _audit(c, x_user, "月份确认", f"{year} 锁定推进：1-{yr['lock_month']}月 → 1-{s.lock_month}月（已确认月改动须走采纳修正）")
    return {"ok": True, "year": year, "lock_month": s.lock_month}


# ---------------- 自然流失预估调参（参考前 n 个月平均） ----------------
class NatParam(BaseModel):
    n: int


@app.post("/api/years/{year}/natparam")
def set_natparam(year: int, p: NatParam, x_user: str = Header("bonniewbli")):
    if not (1 <= p.n <= 24):
        raise HTTPException(422, "回看月数 n 须为 1-24")
    with db() as c:
        require_writer(c, x_user)
        yr = c.execute("SELECT * FROM years WHERE year=?", (year,)).fetchone()
        if not yr:
            raise HTTPException(404, "年份不存在")
        old_n = yr["nat_n"] if "nat_n" in yr.keys() else NAT_N_DEFAULT
        c.execute("UPDATE years SET nat_n=? WHERE year=?", (p.n, year))
        _audit(c, x_user, "自然流失调参", f"{year} 回看月数 n：{old_n} → {p.n}（近{p.n}个月实际离职均值摊到未发生月，重算）")
    return get_board(year)


# ---------------- 分支停用 / 还原（不删数据） ----------------
@app.post("/api/board/{year}/branch/{bid}/toggle")
def branch_toggle(year: int, bid: int, x_user: str = Header("bonniewbli")):
    with db() as c:
        require_writer(c, x_user)
        b = c.execute("SELECT * FROM branches WHERE id=? AND year=?", (bid, year)).fetchone()
        if not b:
            raise HTTPException(404, "分支不存在")
        newv = 0 if b["on_ok"] else 1
        c.execute("UPDATE branches SET on_ok=? WHERE id=?", (newv, bid))
        _audit(c, x_user, "启用分支" if newv else "停用分支",
               f"{year} {b['sec']} · {b['name']}（数据保留" + ("，已恢复计入合计）" if newv else "，可随时还原）"))
    return get_board(year)


# ---------------- 单元格变更历史（口径可复现） ----------------
@app.get("/api/board/{year}/history")
def cell_history(year: int, metric: str, month: int):
    with db() as c:
        rows = [dict(r) for r in c.execute(
            "SELECT * FROM cells_history WHERE year=? AND metric=? AND month=? ORDER BY id DESC LIMIT 50",
            (year, metric, month))]
        return {"year": year, "metric": metric, "metric_name": _metric_name(c, metric), "month": month, "history": rows}


# ---------------- 导出 xlsx ----------------
@app.get("/api/export/{year}.xlsx")
def export_xlsx(year: int):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from fastapi.responses import Response
    board = get_board(year)
    names = {k: n for k, _s, n, *_ in CANON_PROJECTS}
    wb = Workbook()
    ws = wb.active
    ws.title = f"看板1-{year}"
    header = ["项目"] + [f"{m}月" for m in range(1, 13)] + ["年均"]
    ws.append(header)
    for cell in ws[1]:
        cell.font = Font(name="微软雅黑", bold=True)
        cell.fill = PatternFill("solid", fgColor="F3F8FF")
    def row(name, vals, avg_=None):
        ws.append([name] + [("" if v is None else v) for v in vals] + [avg_ if avg_ is not None else ""])
    row(names["budget"], board["metrics"]["budget"]["vals"], board["computed"]["budget_avg"])
    row("总流出（−）", board["computed"]["outT"])
    row("总流入（＋）", board["computed"]["inT"])
    # 260723 口径：实际与预估合一行
    row(names["actual"], board["metrics"]["actual"]["vals"], board["computed"]["chain_avg"])
    ws.column_dimensions["A"].width = 22
    buf = io.BytesIO()
    wb.save(buf)
    with db() as c:
        _audit(c, "system", "导出", f"{year} 看板1 导出 xlsx")
    return Response(buf.getvalue(),
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f"attachment; filename=board1-{year}.xlsx"})


class LedgerEdit(BaseModel):
    fields: dict  # 前端字段名 → 新值（仅人工列；派生列不存库）


@app.post("/api/ledger/row")
def ledger_add_row(e: LedgerEdit, x_user: str = Header("bonniewbli")):
    with db() as c:
        require_writer(c, x_user)
        vals = {db_f: "" for db_f in LEDGER_F2DB.values()}
        for f, v in (e.fields or {}).items():
            if f in LEDGER_F2DB:
                vals[LEDGER_F2DB[f]] = _ledger_norm(f, v)
        missing = [lab for f, lab in LEDGER_REQUIRED if not vals[LEDGER_F2DB[f]]]
        if missing:
            raise HTTPException(422, "必填项未填：" + "、".join(missing))
        if not vals["num"]:
            vals["num"] = "1"  # 每行=一个名额
        try:
            num_ok = float(vals["num"]) == 1
        except ValueError:
            num_ok = False
        if not num_ok:
            raise HTTPException(422, f"招聘数量「{vals['num']}」必须为 1——每行一个名额，多名额请另起一行")
        if vals["st"] and vals["st"] not in LEDGER_ST_CANON:
            raise HTTPException(422, f"当前状态「{vals['st']}」不在下拉枚举（{'/'.join(LEDGER_ST_CANON)}）")
        if vals["cls"] and vals["cls"] not in LEDGER_CLS:
            raise HTTPException(422, "国内/海外 须为：国内 或 海外")
        c.execute(
            "INSERT INTO ledger_rows(year,batch,dept,center,owner,src,job,lvl,fam,cls,loc,ask,num,tgt,st,eta,prev_eta,memo,offer,olvl,join_dt,jmemo,who) "
            "VALUES(0,0,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (vals["dept"], vals["center"], vals["owner"], vals["src"], vals["job"], vals["lvl"], vals["fam"], vals["cls"], vals["loc"],
             vals["ask"], vals["num"], vals["tgt"], vals["st"], vals["eta"], vals["prev_eta"], vals["memo"],
             vals["offer"], vals["olvl"], vals["join_dt"], vals["jmemo"], vals["who"]),
        )
        rid = c.execute("SELECT last_insert_rowid() AS i").fetchone()["i"]
        _audit(c, x_user, "台账新增行", f"行#{rid}" + (f"（{vals['center']} · {vals['job']}）" if vals["center"] or vals["job"] else "（空行·待填）"))
    with db() as c:
        return {"ok": True, "id": rid, "rows": _ledger_list(c, 0)}


@app.put("/api/ledger/row/{rid}")
def ledger_edit_row(rid: int, e: LedgerEdit, x_user: str = Header("bonniewbli")):
    with db() as c:
        require_writer(c, x_user)
        row = c.execute("SELECT * FROM ledger_rows WHERE id=? AND year=0", (rid,)).fetchone()
        if not row:
            raise HTTPException(404, "台账行不存在")
        changes = []
        req = dict(LEDGER_REQUIRED)
        fields = e.fields or {}
        # 列性质闸：数量=1 / 状态枚举 / 国内海外枚举
        if "num" in fields:
            nv_num = _ledger_norm("num", fields["num"]) or "1"
            try:
                if float(nv_num) != 1:
                    raise HTTPException(422, f"招聘数量「{nv_num}」必须为 1——每行一个名额，多名额请另起一行")
            except ValueError:
                raise HTTPException(422, f"招聘数量「{nv_num}」必须为 1——每行一个名额，多名额请另起一行")
        if "st" in fields:
            nv_st = _ledger_norm("st", fields["st"])
            if nv_st and nv_st not in LEDGER_ST_CANON:
                raise HTTPException(422, f"当前状态「{nv_st}」不在下拉枚举（{'/'.join(LEDGER_ST_CANON)}）")
        if "cls" in fields:
            nv_cls = _ledger_norm("cls", fields["cls"])
            if nv_cls and nv_cls not in LEDGER_CLS:
                raise HTTPException(422, "国内/海外 须为：国内 或 海外")
        # 预计到岗变更闸：须填变更原因（memo），旧值自动写入「上次预计到岗(参考)」
        if "eta" in fields:
            nv_eta = _ledger_norm("eta", fields["eta"])
            old_eta = row["eta"] or ""
            if nv_eta != old_eta:
                new_memo = _ledger_norm("memo", fields.get("memo", ""))
                if not new_memo:
                    raise HTTPException(422, "预计到岗时间变更须填写「变更原因/卡点备注」（模板 P 列规则）")
                if old_eta:
                    c.execute("UPDATE ledger_rows SET prev_eta=? WHERE id=?", (old_eta, rid))
                    changes.append(f"peta:自动记录上次预计到岗「{old_eta}」")
        for f, v in fields.items():
            if f not in LEDGER_F2DB:
                continue
            db_f = LEDGER_F2DB[f]
            nv = _ledger_norm(f, v)
            if f in req and not nv:
                raise HTTPException(422, f"必填项「{req[f]}」不能清空")
            ov = row[db_f] or ""
            if nv != ov:
                c.execute(f"UPDATE ledger_rows SET {db_f}=? WHERE id=?", (nv, rid))
                changes.append(f"{f}:「{ov}」→「{nv}」")
        if changes:
            _audit(c, x_user, "台账改行", f"行#{rid} " + "；".join(changes)[:300])
    with db() as c:
        return {"ok": True, "rows": _ledger_list(c, 0)}


@app.delete("/api/ledger/row/{rid}")
def ledger_del_row(rid: int, x_user: str = Header("bonniewbli")):
    with db() as c:
        require_writer(c, x_user)
        row = c.execute("SELECT * FROM ledger_rows WHERE id=? AND year=0", (rid,)).fetchone()
        if not row:
            raise HTTPException(404, "台账行不存在")
        c.execute("DELETE FROM ledger_rows WHERE id=?", (rid,))
        _audit(c, x_user, "台账删行", f"行#{rid}（{row['center']} · {row['job']} · {row['st']}）")
    with db() as c:
        return {"ok": True, "rows": _ledger_list(c, 0)}


@app.get("/api/ledger")
def get_ledger():
    with db() as c:
        return {"rows": _ledger_list(c, 0)}


@app.post("/api/ledger/import")
async def import_ledger(file: UploadFile = File(...), x_user: str = Header("bonniewbli")):
    """台账全局一份（跨年滚动）；3.2 矩阵由前端按归月年份分发到各年份页签"""
    year = 0
    with db() as c:
        require_writer(c, x_user)
    raw = await file.read()
    rows, report = parse_ledger(raw, file.filename or "upload")
    if not rows:
        raise HTTPException(422, "识别到表头但无有效数据行")
    errors = validate_ledger_rows(rows)  # 列性质闸：枚举/日期limit/数量=1，不合格整批拒绝
    if errors:
        raise HTTPException(422, {"msg": f"台账数据不合格 {len(errors)} 处，整批拒绝入库（一处都不能错）",
                                  "errors": errors[:50], "total_errors": len(errors)})
    with db() as c:
        require_writer(c, x_user)
        c.execute(
            "INSERT INTO ledger_snapshots(year,filename,sheet,rows_n,created_by,created_at) VALUES(?,?,?,?,?,?)",
            (year, file.filename, report["sheet"], len(rows), x_user, now()),
        )
        batch = c.execute("SELECT last_insert_rowid() AS i").fetchone()["i"]
        c.execute("DELETE FROM ledger_rows WHERE year=?", (year,))
        for r in rows:
            c.execute(
                "INSERT INTO ledger_rows(year,batch,dept,center,owner,src,job,lvl,fam,cls,loc,ask,num,tgt,st,eta,prev_eta,memo,offer,olvl,join_dt,jmemo,who) "
                "VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                (year, batch, r["dept"], r["center"], r["owner"], r["src"], r["job"], r["lvl"], r["fam"], r["cls"], r["loc"],
                 r["ask"], r["num"], r["tgt"], r["st"], r["eta"], r["prev_eta"], r["memo"], r["offer"], r["olvl"], r["join"], r["jmemo"], r["who"]),
            )
        _audit(c, x_user, "导入台账",
               f"全局台账 批次#{batch}「{file.filename}」sheet「{report['sheet']}」表头第{report['header_row']}行 · "
               f"列名命中{report['hits']}项 · 有效{len(rows)}行/跳过{report['skipped']}行（整年替换）")
    with db() as c:
        return {"ok": True, "report": report, "rows": _ledger_list(c, year)}


# ---------------- 台账模板下载（与线下模板 v2 同构：分组表头/下拉/冻结/说明行） ----------------
@app.get("/api/ledger/template.xlsx")
def ledger_template():
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation
    from fastapi.responses import Response
    wb = Workbook()
    ws = wb.active
    ws.title = "台账"
    groups = [("需求信息", 12), ("进展&状态机", 5), ("入选与入职", 4), ("系统派生（自动·勿填）", 5)]
    headers = ["部门", "中心", "业务负责人", "HC来源备注", "招聘岗位", "职级", "分类", "国内/海外", "城市",
               "需求提出时间", "招聘数量", "目标到岗时间(初始固定)",
               "当前状态", "预计到岗时间", "上次预计到岗(参考)", "变更原因/卡点备注", "面试中人选(可多条)",
               "offer人选", "人选职级", "实际入职时间", "入职备注",
               "归月(自动)", "计入看板1(自动)", "招聘周期·天(自动)", "超期(自动)", "一句话备注(自动)"]
    hints = ["必填", "必填", "必填", "离职补录/新增投入等", "必填", "如T9", "如研发/产品", "下拉:国内/海外", "如深圳",
             "日期 2026-03-15", "必须为1(多名额另起一行)", "立项时定,永不改",
             "下拉(必填)", "日期;变更须填变更原因", "系统自动记录,勿填", "eta变更时必填", "可多条",
             "", "", "日期", "",
             "勿填", "勿填", "勿填", "勿填", "勿填"]
    col = 1
    for gname, span in groups:
        ws.cell(1, col, gname)
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + span - 1)
        col += span
    ws.append(headers)
    ws.append(hints)
    for cell in ws[1]:
        cell.font = Font(name="微软雅黑", bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0A2E6E")
        cell.alignment = Alignment(horizontal="center")
    for cell in ws[2]:
        cell.font = Font(name="微软雅黑", bold=True)
        cell.fill = PatternFill("solid", fgColor="DBEAFE")
    for cell in ws[3]:
        cell.font = Font(name="微软雅黑", size=9, color="8194AE")
    dv_cls = DataValidation(type="list", formula1='"%s"' % ",".join(LEDGER_CLS), allow_blank=True,
                            errorTitle="无效值", error="须为：国内 或 海外", showErrorMessage=True)
    dv_st = DataValidation(type="list", formula1='"%s"' % ",".join(LEDGER_ST_CANON), allow_blank=True,
                           errorTitle="无效值", error="须从下拉选择", showErrorMessage=True)
    dv_num = DataValidation(type="whole", operator="equal", formula1="1",
                            errorTitle="招聘数量必须为1", error="每行一个名额，多名额请另起一行", showErrorMessage=True)
    ws.add_data_validation(dv_cls)
    ws.add_data_validation(dv_st)
    ws.add_data_validation(dv_num)
    dv_cls.add("H4:H500")
    dv_st.add("M4:M500")
    dv_num.add("K4:K500")
    widths = [10, 10, 12, 14, 18, 7, 9, 11, 8, 13, 12, 16, 13, 13, 14, 16, 16, 12, 9, 13, 10, 10, 14, 12, 9, 22]
    for i, w in enumerate(widths):
        ws.column_dimensions[chr(65 + i) if i < 26 else "A" + chr(65 + i - 26)].width = w
    ws.freeze_panes = "F4"  # 冻结：部门~招聘岗位（A-E）+ 三行表头
    buf = io.BytesIO()
    wb.save(buf)
    return Response(buf.getvalue(),
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": "attachment; filename=ledger-template.xlsx"})


# ---------------- 静态前端（同源托管 index.html / admin.html） ----------------
@app.get("/")
def root():
    return FileResponse(os.path.join(FRONT_DIR, "index.html"))


@app.get("/{page}.html")
def page(page: str):
    fp = os.path.join(FRONT_DIR, f"{page}.html")
    if page in ("index", "admin") and os.path.exists(fp):
        return FileResponse(fp)
    raise HTTPException(404)
